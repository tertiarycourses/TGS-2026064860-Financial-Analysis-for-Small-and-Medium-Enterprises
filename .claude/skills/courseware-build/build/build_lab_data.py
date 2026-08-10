#!/usr/bin/env python3
"""Generate the realistic mock-data Excel workbooks for the hands-on activities
(activities/data/) and the PP assessment data workbook (assessment/).

The AGGREGATE figures are exactly the ones used in the activity model answers and
the PP answer key, so every worked answer in the LG / assessment keys stays valid.
Realism is added *below* the aggregates: named SME companies, line-item detail that
sums to the given totals, SGD formatting, proper statement layout and a workings
template sheet in every workbook. All companies and figures are fictitious.
"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
def _find_repo(start):
    env = os.environ.get("COURSE_REPO")
    if env and os.path.isdir(env):
        return env
    d = start
    for _ in range(8):
        d = os.path.dirname(d)
        if os.path.isdir(os.path.join(d, "courseware")) and os.path.isdir(os.path.join(d, "activities")):
            return d
    return os.path.dirname(os.path.dirname(HERE))
REPO = _find_repo(HERE)
ASSESS_DIR = os.path.join(REPO, "assessment")

def actdir(n):
    """activities/activityNN — one subfolder per activity (md + xlsx + csv)."""
    d = os.path.join(REPO, "activities", f"activity{n:02d}")
    os.makedirs(d, exist_ok=True)
    return d

import csv as _csv
def write_csv(folder, name, header, rows):
    path = os.path.join(folder, name)
    with open(path, "w", newline="") as f:
        w = _csv.writer(f)
        w.writerow(header)
        for row in rows:
            w.writerow(row)
    print("saved", path)

# ---------------- styling helpers ----------------
BLUE = "1F6FEB"; LIGHT = "EAF1FC"; INK = "161B26"; GREY = "5B6372"
thin = Side(style="thin", color="C9D4E4")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

def sheet_title(ws, company, subtitle, note, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=company)
    c.font = Font(name="Arial", size=14, bold=True, color=BLUE)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name="Arial", size=11, bold=True, color=INK)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1, value=note)
    c.font = Font(name="Arial", size=9, italic=True, color=GREY)
    return 5   # first data row

def header_row(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal=("left" if i == 1 else "right"))
        c.border = BOX
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1

def put(ws, row, label, vals, kind="item", numfmt="#,##0", indent=1):
    """kind: item | sub (section header) | total | blank"""
    c = ws.cell(row=row, column=1, value=label)
    c.border = BOX
    if kind == "sub":
        c.font = Font(name="Arial", size=10, bold=True, color=BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        for i in range(2, 2 + len(vals)):
            cc = ws.cell(row=row, column=i)
            cc.fill = PatternFill("solid", fgColor=LIGHT); cc.border = BOX
        return row + 1
    c.font = Font(name="Arial", size=10, bold=(kind == "total"), color=INK)
    c.alignment = Alignment(indent=(0 if kind == "total" else indent))
    if kind == "total":
        c.fill = PatternFill("solid", fgColor="F0F4FA")
    for i, v in enumerate(vals, 2):
        cc = ws.cell(row=row, column=i, value=v)
        cc.number_format = numfmt
        cc.font = Font(name="Arial", size=10, bold=(kind == "total"), color=INK)
        cc.alignment = Alignment(horizontal="right")
        cc.border = BOX
        if kind == "total":
            cc.fill = PatternFill("solid", fgColor="F0F4FA")
    return row + 1

def footer_note(ws, row, ncols):
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=ncols)
    c = ws.cell(row=row + 1, column=1,
                value="Fictitious data prepared for training use only — WSQ Financial Analysis for SMEs (TGS-2026064860), Tertiary Infotech Academy Pte Ltd.")
    c.font = Font(name="Arial", size=8, italic=True, color=GREY)

def workings_sheet(wb, title, lines):
    ws = wb.create_sheet("Your Workings")
    r = sheet_title(ws, title, "Your Workings", "Complete your calculations in this sheet.", 6)
    ws.column_dimensions["A"].width = 46
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 14
    for ln in lines:
        c = ws.cell(row=r, column=1, value=ln)
        c.font = Font(name="Arial", size=10, bold=ln.endswith(":"), color=(BLUE if ln.endswith(":") else INK))
        r += 1
    return ws

# ================================================================ Activity 1
# Havenwood Trading Pte Ltd — figures in S$ million (aggregates = activity data)
def build_a1():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Balance Sheet")
    r = sheet_title(ws, "Havenwood Trading Pte Ltd",
                    "Statement of Financial Position as at 31 December",
                    "All figures in S$ million. Prepared under SFRS for Small Entities.", 3)
    r = header_row(ws, r, ["", "FY2019", "FY2018"], widths=[42, 12, 12])
    r = put(ws, r, "ASSETS", [None, None], kind="sub")
    r = put(ws, r, "Property, plant & equipment (net)", [100, 90])
    r = put(ws, r, "Trade receivables", [50, 50])
    r = put(ws, r, "Inventories", [30, 30])
    r = put(ws, r, "Cash and bank balances", [20, 20])
    r = put(ws, r, "Total assets", [200, 190], kind="total")
    r = put(ws, r, "EQUITY AND LIABILITIES", [None, None], kind="sub")
    r = put(ws, r, "Share capital and retained earnings", [80, 80])
    r = put(ws, r, "Bank loans", [50, 55])
    r = put(ws, r, "Trade payables", [40, 35])
    r = put(ws, r, "Provisions", [30, 20])
    r = put(ws, r, "Total equity and liabilities", [200, 190], kind="total")
    footer_note(ws, r, 3)

    ws = wb.create_sheet("Income Statement")
    r = sheet_title(ws, "Havenwood Trading Pte Ltd",
                    "Income Statement for the year ended 31 December 2019",
                    "All figures in S$ million.", 2)
    r = header_row(ws, r, ["", "FY2019"], widths=[42, 12])
    r = put(ws, r, "Revenue", [None], kind="sub")
    r = put(ws, r, "Wholesale distribution", [38])
    r = put(ws, r, "Retail and e-commerce", [12])
    r = put(ws, r, "Total revenue", [50], kind="total")
    r = put(ws, r, "Cost of goods sold", [10])
    r = put(ws, r, "Operational costs", [None], kind="sub")
    r = put(ws, r, "Salaries and CPF", [12])
    r = put(ws, r, "Rental of premises", [4])
    r = put(ws, r, "Utilities, logistics and admin", [4])
    r = put(ws, r, "Total operational costs", [20], kind="total")
    r = put(ws, r, "EBITDA", [20], kind="total")
    r = put(ws, r, "Depreciation", [5])
    r = put(ws, r, "Interest cost", [3])
    r = put(ws, r, "Tax", [2])
    r = put(ws, r, "Net income", [10], kind="total")
    footer_note(ws, r, 2)

    workings_sheet(wb, "Havenwood Trading Pte Ltd", [
        "Operating Cash Flow:", "EBITDA", "Tax", "Change in net working capital", "Change in provisions",
        "Operating cash flow (total)", "",
        "Investing Cash Flow:", "Change in fixed assets", "Depreciation 2019", "Change in payables on capex",
        "Investing cash flow (total)", "",
        "Financing Cash Flow:", "Change in financial debt", "Interest cost 2019", "Change in earnings",
        "Financing cash flow (total)", "", "Net cash flow:"])
    d = actdir(1)
    out = os.path.join(d, "FA-Activity-01-Cash-Flow-Statement.xlsx")
    wb.save(out); print("saved", out)
    write_csv(d, "balance_sheet.csv", ["Item", "FY2019", "FY2018"], [
        ["Property, plant & equipment (net)", 100, 90], ["Trade receivables", 50, 50],
        ["Inventories", 30, 30], ["Cash and bank balances", 20, 20], ["Total assets", 200, 190],
        ["Share capital and retained earnings", 80, 80], ["Bank loans", 50, 55],
        ["Trade payables", 40, 35], ["Provisions", 30, 20], ["Total equity and liabilities", 200, 190]])
    write_csv(d, "income_statement.csv", ["Item", "FY2019"], [
        ["Wholesale distribution revenue", 38], ["Retail and e-commerce revenue", 12],
        ["Total revenue", 50], ["Cost of goods sold", 10], ["Salaries and CPF", 12],
        ["Rental of premises", 4], ["Utilities, logistics and admin", 4], ["Total operational costs", 20],
        ["EBITDA", 20], ["Depreciation", 5], ["Interest cost", 3], ["Tax", 2], ["Net income", 10]])

# ================================================================ Activities 2 & 6
# Company X = Orchid Logistics Pte Ltd · Company Y = Marina Retail Group Pte Ltd
def _xy_statements(wb):
    ws = wb.create_sheet("Balance Sheet")
    r = sheet_title(ws, "Orchid Logistics Pte Ltd (Company X)  ·  Marina Retail Group Pte Ltd (Company Y)",
                    "Statements of Financial Position as at 31 December 2025",
                    "All figures in S$. Two fictitious SMEs prepared for comparison.", 3)
    r = header_row(ws, r, ["", "Company X", "Company Y"], widths=[42, 14, 14])
    r = put(ws, r, "CURRENT ASSETS", [None, None], kind="sub")
    r = put(ws, r, "Cash and bank balances", [50000, 80000])
    r = put(ws, r, "Inventories", [40000, 70000])
    r = put(ws, r, "Trade receivables", [100000, 250000])
    r = put(ws, r, "FIXED ASSETS", [None, None], kind="sub")
    r = put(ws, r, "Property", [400000, 450000])
    r = put(ws, r, "Other fixed assets", [60000, 50000])
    r = put(ws, r, "Total assets", [650000, 900000], kind="total")
    r = put(ws, r, "LIABILITIES AND EQUITY", [None, None], kind="sub")
    r = put(ws, r, "Trade payables", [100000, 150000])
    r = put(ws, r, "Short-term loans", [80000, 70000])
    r = put(ws, r, "Long-term liabilities", [300000, 500000])
    r = put(ws, r, "Capital (equity)", [170000, 180000])
    r = put(ws, r, "Total liabilities and equity", [650000, 900000], kind="total")
    footer_note(ws, r, 3)

    ws = wb.create_sheet("Income Statement")
    r = sheet_title(ws, "Orchid Logistics Pte Ltd (X)  ·  Marina Retail Group Pte Ltd (Y)",
                    "Income Statements for the year ended 31 December 2025",
                    "All figures in S$.", 3)
    r = header_row(ws, r, ["", "Company X", "Company Y"], widths=[42, 14, 14])
    r = put(ws, r, "Sales", [300000, 500000])
    r = put(ws, r, "Cost of goods sold", [140000, 200000])
    r = put(ws, r, "Gross profit", [160000, 300000], kind="total")
    r = put(ws, r, "Selling and distribution expenses", [50000, 60000])
    r = put(ws, r, "EBITDA", [110000, 240000], kind="total")
    r = put(ws, r, "Depreciation", [10000, 12000])
    r = put(ws, r, "EBIT", [100000, 228000], kind="total")
    r = put(ws, r, "Interest", [10000, 18000])
    r = put(ws, r, "Earnings before tax", [90000, 210000], kind="total")
    r = put(ws, r, "Taxes", [18000, 42000])
    r = put(ws, r, "Net income", [72000, 168000], kind="total")
    footer_note(ws, r, 3)

XY_BS_ROWS = [
    ["Cash and bank balances", 50000, 80000], ["Inventories", 40000, 70000],
    ["Trade receivables", 100000, 250000], ["Property", 400000, 450000],
    ["Other fixed assets", 60000, 50000], ["Total assets", 650000, 900000],
    ["Trade payables", 100000, 150000], ["Short-term loans", 80000, 70000],
    ["Long-term liabilities", 300000, 500000], ["Capital (equity)", 170000, 180000],
    ["Total liabilities and equity", 650000, 900000]]
XY_IS_ROWS = [
    ["Sales", 300000, 500000], ["Cost of goods sold", 140000, 200000],
    ["Gross profit", 160000, 300000], ["Selling and distribution expenses", 50000, 60000],
    ["EBITDA", 110000, 240000], ["Depreciation", 10000, 12000], ["EBIT", 100000, 228000],
    ["Interest", 10000, 18000], ["Earnings before tax", 90000, 210000],
    ["Taxes", 18000, 42000], ["Net income", 72000, 168000]]

def build_a2():
    wb = Workbook(); wb.remove(wb.active)
    _xy_statements(wb)
    workings_sheet(wb, "Ratio Analysis — Company X vs Company Y", [
        "Liquidity ratios:", "Current ratio", "Quick ratio", "",
        "Profitability ratios:", "Operating profit ratio", "Net profit ratio", "Return on capital employed", "",
        "Turnover ratios (assume purchases = 50% of COGS):",
        "Inventory turnover", "Receivables turnover", "Payables turnover", "",
        "Solvency ratios:", "Debt-to-equity ratio", "Financial leverage", "",
        "Verdict per ratio family:"])
    d = actdir(2)
    out = os.path.join(d, "FA-Activity-02-Ratio-Analysis.xlsx")
    wb.save(out); print("saved", out)
    write_csv(d, "balance_sheet_x_y.csv", ["Item", "Company X", "Company Y"], XY_BS_ROWS)
    write_csv(d, "income_statement_x_y.csv", ["Item", "Company X", "Company Y"], XY_IS_ROWS)

def build_a6():
    wb = Workbook(); wb.remove(wb.active)
    _xy_statements(wb)
    workings_sheet(wb, "Solvency & Financial Risk — Company X vs Company Y", [
        "Debt-to-equity ratio (LT debt / equity):",
        "Financial leverage (total assets / equity):",
        "Interest coverage ratio (EBIT / interest):", "",
        "Risk verdict:"])
    d = actdir(6)
    out = os.path.join(d, "FA-Activity-06-Solvency-Analysis.xlsx")
    wb.save(out); print("saved", out)
    write_csv(d, "balance_sheet_x_y.csv", ["Item", "Company X", "Company Y"], XY_BS_ROWS)
    write_csv(d, "income_statement_x_y.csv", ["Item", "Company X", "Company Y"], XY_IS_ROWS)

# ================================================================ Activity 3
# Sunrise F&B Holdings Pte Ltd — 4-year trend (aggregates = activity data)
def build_a3():
    Y = ["FY2021", "FY2020", "FY2019", "FY2018"]
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Balance Sheet 2018-2021")
    r = sheet_title(ws, "Sunrise F&B Holdings Pte Ltd",
                    "Statements of Financial Position as at 31 December (4 years)",
                    "All figures in S$.", 5)
    r = header_row(ws, r, [""] + Y, widths=[42, 13, 13, 13, 13])
    r = put(ws, r, "CURRENT ASSETS", [None]*4, kind="sub")
    r = put(ws, r, "Cash and bank balances", [40000, 40000, 40000, 40000])
    r = put(ws, r, "Inventories", [40000, 34000, 32000, 45000])
    r = put(ws, r, "Trade receivables", [75583, 69984, 64800, 60000])
    r = put(ws, r, "FIXED ASSETS", [None]*4, kind="sub")
    r = put(ws, r, "Property", [144000, 146000, 148000, 150000])
    r = put(ws, r, "Other fixed assets", [39998, 39999, 40000, 35000])
    r = put(ws, r, "Total assets", [339581, 329983, 324800, 330000], kind="total")
    r = put(ws, r, "LIABILITIES AND EQUITY", [None]*4, kind="sub")
    r = put(ws, r, "Trade payables", [75000, 60000, 52000, 43000])
    r = put(ws, r, "Short-term loans", [20000, 32000, 27000, 25000])
    r = put(ws, r, "Long-term liabilities", [50000, 50000, 50000, 50000])
    r = put(ws, r, "Capital and retained earnings", [194581, 187983, 195800, 212000])
    r = put(ws, r, "Total liabilities and equity", [339581, 329983, 324800, 330000], kind="total")
    footer_note(ws, r, 5)

    ws = wb.create_sheet("Income Statement 2018-2021")
    r = sheet_title(ws, "Sunrise F&B Holdings Pte Ltd",
                    "Income Statements for the years ended 31 December (4 years)",
                    "All figures in S$.", 5)
    r = header_row(ws, r, [""] + Y, widths=[42, 13, 13, 13, 13])
    r = put(ws, r, "Sales", [239580, 217800, 198000, 180000])
    r = put(ws, r, "Cost of goods sold", [98398, 93713, 89250, 85000])
    r = put(ws, r, "Gross profit", [141182, 124088, 108750, 95000], kind="total")
    r = put(ws, r, "Selling and distribution expenses", [20000, 20000, 20000, 20000])
    r = put(ws, r, "EBITDA", [121182, 104088, 88750, 75000], kind="total")
    r = put(ws, r, "Depreciation", [3000, 3000, 3000, 3000])
    r = put(ws, r, "EBIT", [118182, 101088, 85750, 72000], kind="total")
    r = put(ws, r, "Interest", [4997, 4998, 4999, 5000])
    r = put(ws, r, "Earnings before tax", [113185, 96090, 80751, 67000], kind="total")
    r = put(ws, r, "Taxes", [22637, 19218, 16150, 13400])
    r = put(ws, r, "Net income", [90548, 76872, 64601, 53600], kind="total")
    footer_note(ws, r, 5)

    ws = wb.create_sheet("Trend Template")
    r = sheet_title(ws, "Sunrise F&B Holdings Pte Ltd",
                    "Trend-Analysis Template — fill in the YoY growth and the ratios",
                    "YoY growth = (this year − last year) / last year.", 8)
    r = header_row(ws, r, ["Ratio", "FY2021", "YoY %", "FY2020", "YoY %", "FY2019", "YoY %", "FY2018"],
                   widths=[34, 11, 9, 11, 9, 11, 9, 11])
    for section, items in [
        ("Liquidity ratios", ["Current ratio", "Acid (quick) ratio"]),
        ("Efficiency ratios (purchases = 50% of COGS)",
         ["Asset turnover", "Inventory turnover", "Payables turnover", "Receivables turnover"]),
        ("Leverage ratios", ["Debt-to-equity ratio", "Debt-to-assets ratio"]),
        ("Profitability ratios", ["Operating margin", "Net profit margin"]),
    ]:
        r = put(ws, r, section, [None]*7, kind="sub")
        for it in items:
            r = put(ws, r, it, [None]*7)
    footer_note(ws, r, 8)
    d = actdir(3)
    out = os.path.join(d, "FA-Activity-03-Trend-Analysis.xlsx")
    wb.save(out); print("saved", out)
    write_csv(d, "balance_sheet_2018_2021.csv", ["Item", "FY2021", "FY2020", "FY2019", "FY2018"], [
        ["Cash and bank balances", 40000, 40000, 40000, 40000],
        ["Inventories", 40000, 34000, 32000, 45000],
        ["Trade receivables", 75583, 69984, 64800, 60000],
        ["Property", 144000, 146000, 148000, 150000],
        ["Other fixed assets", 39998, 39999, 40000, 35000],
        ["Total assets", 339581, 329983, 324800, 330000],
        ["Trade payables", 75000, 60000, 52000, 43000],
        ["Short-term loans", 20000, 32000, 27000, 25000],
        ["Long-term liabilities", 50000, 50000, 50000, 50000],
        ["Capital and retained earnings", 194581, 187983, 195800, 212000],
        ["Total liabilities and equity", 339581, 329983, 324800, 330000]])
    write_csv(d, "income_statement_2018_2021.csv", ["Item", "FY2021", "FY2020", "FY2019", "FY2018"], [
        ["Sales", 239580, 217800, 198000, 180000],
        ["Cost of goods sold", 98398, 93713, 89250, 85000],
        ["Gross profit", 141182, 124088, 108750, 95000],
        ["Selling and distribution expenses", 20000, 20000, 20000, 20000],
        ["EBITDA", 121182, 104088, 88750, 75000],
        ["Depreciation", 3000, 3000, 3000, 3000],
        ["EBIT", 118182, 101088, 85750, 72000],
        ["Interest", 4997, 4998, 4999, 5000],
        ["Earnings before tax", 113185, 96090, 80751, 67000],
        ["Taxes", 22637, 19218, 16150, 13400],
        ["Net income", 90548, 76872, 64601, 53600]])

# ================================================================ Activities 4 & 5
# Refrigerated delivery van proposal — S$'000 (aggregates = activity data)
def _project_sheet(wb, extra_cols, template_title):
    ws = wb.create_sheet("Project Cash Flows")
    r = sheet_title(ws, "Havenwood Trading Pte Ltd — Investment Proposal",
                    "Refrigerated delivery van: purchase vs outsourced cold-chain delivery",
                    "All figures in S$'000. Outlay S$2,000k in Year 0; net cash inflow S$375k per year for 10 years; required return 10%.", 2 + len(extra_cols))
    r = header_row(ws, r, ["Year", "Net cash flow"] + extra_cols,
                   widths=[10, 14] + [16]*len(extra_cols))
    r = put(ws, r, "0", [-2000] + [None]*len(extra_cols))
    for y in range(1, 11):
        r = put(ws, r, str(y), [375] + [None]*len(extra_cols))
    footer_note(ws, r, 2 + len(extra_cols))
    return ws

def build_a4():
    wb = Workbook(); wb.remove(wb.active)
    _project_sheet(wb, ["Cumulative cash flow", "Discounted cash flow (10%)", "Cumulative discounted"],
                   "Payback template")
    workings_sheet(wb, "Payback Period & Discounted Payback", [
        "Simple payback period (years):",
        "Discounted payback period (years):", "",
        "Discount factor: 1 / (1.10)^n", "Why does discounting lengthen the payback?"])
    d = actdir(4)
    out = os.path.join(d, "FA-Activity-04-Payback-Period.xlsx")
    wb.save(out); print("saved", out)
    write_csv(d, "project_cash_flows.csv", ["Year", "Net cash flow (S$'000)"],
              [[0, -2000]] + [[y, 375] for y in range(1, 11)])

def build_a5():
    wb = Workbook(); wb.remove(wb.active)
    _project_sheet(wb, ["Discounted cash flow (10%)"], "NPV / PI template")
    workings_sheet(wb, "Net Present Value & Profitability Index", [
        "PV of cash inflows:", "PV of cash outflows:", "Net Present Value (NPV):",
        "Profitability Index (PI):", "", "Decision (accept / reject) and reason:"])
    d = actdir(5)
    out = os.path.join(d, "FA-Activity-05-NPV-PI.xlsx")
    wb.save(out); print("saved", out)
    write_csv(d, "project_cash_flows.csv", ["Year", "Net cash flow (S$'000)"],
              [[0, -2000]] + [[y, 375] for y in range(1, 11)])

# ================================================================ PP assessment data
# Company A = Kestrel Precision Engineering · Company B = Lighthouse Marine Supplies
# Year 3 of Company A matches the PP answer key exactly:
#   Sales 140,000 · Net income 36,000 · Total assets 310,000 · Equity 135,000
#   → NPM 25.71% · TAT 45.17% · EM 2.30 · ROE 26.71%
# Company A performs better than B on ROE/ROA/NPM in every year (per the key's verdict).
def build_pp_data():
    Y = ["Year 3 (FY2025)", "Year 2 (FY2024)", "Year 1 (FY2023)"]
    A_BS = dict(cash=[28000, 24000, 21000], inv=[30000, 28000, 27000], ar=[42000, 40000, 38000],
                prop=[180000, 176000, 172000], other=[30000, 28000, 26000],
                ap=[45000, 42000, 40000], stl=[30000, 30000, 30000], ltl=[100000, 105000, 110000],
                eq=[135000, 119000, 104000])
    B_BS = dict(cash=[30000, 32000, 30000], inv=[60000, 56000, 52000], ar=[85000, 78000, 72000],
                prop=[210000, 208000, 206000], other=[35000, 34000, 33000],
                ap=[70000, 66000, 62000], stl=[45000, 45000, 45000], ltl=[185000, 190000, 195000],
                eq=[120000, 107000, 91000])
    #  Net income is fixed by the answer key (A Year 3: 36,000 on sales 140,000); the tax
    #  line is computed as EBT − NI so every statement casts correctly (~17-19% effective).
    A_IS = dict(sales=[140000, 128000, 118000], cogs=[55000, 53000, 50000],
                sd=[26000, 25500, 24500], dep=[9000, 9000, 9000],
                interest=[5600, 6000, 6300], ni=[36000, 28000, 22000])
    B_IS = dict(sales=[190000, 180000, 172000], cogs=[105000, 100000, 96000],
                sd=[42000, 40000, 39000], dep=[11000, 11000, 11000],
                interest=[11500, 11800, 12100], ni=[17000, 14000, 11500])

    def totals(bs):
        return [bs["cash"][i] + bs["inv"][i] + bs["ar"][i] + bs["prop"][i] + bs["other"][i] for i in range(3)]

    wb = Workbook(); wb.remove(wb.active)
    for name, bs, is_ in [("Company A — Kestrel Precision Engineering Pte Ltd", A_BS, A_IS),
                          ("Company B — Lighthouse Marine Supplies Pte Ltd", B_BS, B_IS)]:
        short = "Company A" if name.startswith("Company A") else "Company B"
        ws = wb.create_sheet(f"{short} Balance Sheet")
        r = sheet_title(ws, name, "Statements of Financial Position as at 31 December (3 years)",
                        "All figures in S$.", 4)
        r = header_row(ws, r, [""] + Y, widths=[42, 15, 15, 15])
        r = put(ws, r, "CURRENT ASSETS", [None]*3, kind="sub")
        r = put(ws, r, "Cash and bank balances", bs["cash"])
        r = put(ws, r, "Inventories", bs["inv"])
        r = put(ws, r, "Trade receivables", bs["ar"])
        r = put(ws, r, "FIXED ASSETS", [None]*3, kind="sub")
        r = put(ws, r, "Property, plant & equipment", bs["prop"])
        r = put(ws, r, "Other fixed assets", bs["other"])
        r = put(ws, r, "Total assets", totals(bs), kind="total")
        r = put(ws, r, "LIABILITIES AND EQUITY", [None]*3, kind="sub")
        r = put(ws, r, "Trade payables", bs["ap"])
        r = put(ws, r, "Short-term loans", bs["stl"])
        r = put(ws, r, "Long-term liabilities", bs["ltl"])
        r = put(ws, r, "Capital and retained earnings", bs["eq"])
        tot = [bs["ap"][i] + bs["stl"][i] + bs["ltl"][i] + bs["eq"][i] for i in range(3)]
        r = put(ws, r, "Total liabilities and equity", tot, kind="total")
        footer_note(ws, r, 4)

        ws = wb.create_sheet(f"{short} Income Statement")
        r = sheet_title(ws, name, "Income Statements for the years ended 31 December (3 years)",
                        "All figures in S$.", 4)
        r = header_row(ws, r, [""] + Y, widths=[42, 15, 15, 15])
        gp = [is_["sales"][i] - is_["cogs"][i] for i in range(3)]
        ebitda = [gp[i] - is_["sd"][i] for i in range(3)]
        ebit = [ebitda[i] - is_["dep"][i] for i in range(3)]
        ebt = [ebit[i] - is_["interest"][i] for i in range(3)]
        tax = [ebt[i] - is_["ni"][i] for i in range(3)]
        r = put(ws, r, "Sales", is_["sales"])
        r = put(ws, r, "Cost of goods sold", is_["cogs"])
        r = put(ws, r, "Gross profit", gp, kind="total")
        r = put(ws, r, "Selling and distribution expenses", is_["sd"])
        r = put(ws, r, "EBITDA", ebitda, kind="total")
        r = put(ws, r, "Depreciation", is_["dep"])
        r = put(ws, r, "EBIT", ebit, kind="total")
        r = put(ws, r, "Interest", is_["interest"])
        r = put(ws, r, "Earnings before tax", ebt, kind="total")
        r = put(ws, r, "Taxes", tax)
        r = put(ws, r, "Net income", is_["ni"], kind="total")
        footer_note(ws, r, 4)

    workings_sheet(wb, "PP Assessment — Your Analysis", [
        "Task 1 (A1) — ratio trends for Company A and Company B:",
        "ROE · ROA · Net profit margin · Current ratio · Debt-to-equity (per year)", "",
        "Task 2 (A2) — DuPont analysis for Company A (Year 3):",
        "Net profit margin = Net income / Sales",
        "Asset turnover = Sales / Total assets",
        "Equity multiplier = Total assets / Equity",
        "ROE = NPM × Asset turnover × Equity multiplier"])
    out = os.path.join(ASSESS_DIR, "PP Data - Financial Analysis for Small and Medium Enterprises - Company A and B.xlsx")
    wb.save(out); print("saved", out)
    # sanity: year-3 DuPont for Company A must match the answer key
    assert A_IS["ni"][0] == 36000 and A_IS["sales"][0] == 140000
    assert totals(A_BS)[0] == 310000, totals(A_BS)
    assert A_BS["eq"][0] == 135000

if __name__ == "__main__":
    build_a1(); build_a2(); build_a3(); build_a4(); build_a5(); build_a6(); build_pp_data()
    print("All mock-data workbooks generated.")
